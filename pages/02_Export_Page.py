# export_page.py - Excel Export Functionality for Medical Schedule Management
# Generate Excel files matching the template format with dynamic session columns

import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import calendar
from typing import Dict, List, Any, Tuple
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from config import (
    MORNING_CUTOFF_TIME,
    OFF_SESSION_PLACEHOLDER,
    POC_TEAM_LEADER,
    POC_DOCTORS,
    WEEKEND_DAYS,
)
from data_ops import load_submissions, format_doctor_name, format_doctor_first_last_name
from validation import is_weekend, calculate_total_duration

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# Set up Streamlit page configuration
st.set_page_config(
    page_title="Export To Excel",
    # page_icon=APP_ICON,
    layout="wide",
    initial_sidebar_state="collapsed",
)

# =============================================================================
# DATA PROCESSING FUNCTIONS
# =============================================================================


def get_month_data(team_leader_id: str, month: str) -> Dict[str, List[Dict]]:
    """
    Get all submissions data for a specific team leader and month.

    Args:
        team_leader_id: Team leader ID
        month: Month in YYYY-MM format

    Returns:
        Dictionary with doctor_id as keys and list of sessions as values
    """
    logger.info(f"Loading month data for TL: {team_leader_id}, Month: {month}")

    all_submissions = load_submissions()
    logger.info(f"Loaded {len(all_submissions)} total submissions")

    if not all_submissions:
        logger.warning("No submissions found")
        return {}

    # Filter by month
    month_start = datetime.strptime(month + "-01", "%Y-%m-%d")
    month_end = month_start.replace(
        month=month_start.month % 12 + 1, day=1
    ) - timedelta(days=1)

    logger.info(f"Filtering for date range: {month_start.date()} to {month_end.date()}")

    # Group sessions by doctor
    doctor_sessions = {}
    filtered_count = 0

    for submission in all_submissions:
        try:
            submission_date = datetime.strptime(submission["date"], "%Y-%m-%d")

            # Check if submission is in the target month
            if month_start.date() <= submission_date.date() <= month_end.date():
                doctor_id = submission["doctor_id"]

                if doctor_id not in doctor_sessions:
                    doctor_sessions[doctor_id] = []

                doctor_sessions[doctor_id].append(submission)
                filtered_count += 1
        except ValueError as e:
            logger.error(
                f"Invalid date format in submission: {submission.get('date')} - {e}"
            )
            continue

    logger.info(f"Filtered to {filtered_count} submissions for the target month")
    logger.info(
        f"Sessions by doctor: {[(doc_id, len(sessions)) for doc_id, sessions in doctor_sessions.items()]}"
    )

    return doctor_sessions


def calculate_max_sessions_per_day(sessions: List[Dict]) -> int:
    """
    Calculate the maximum number of sessions per day for a given set of sessions.

    Args:
        sessions: List of session dictionaries

    Returns:
        Maximum number of sessions in a single day
    """
    if not sessions:
        return 1  # Minimum 1 session column

    # Group sessions by date
    sessions_by_date = {}
    for session in sessions:
        date_key = session["date"]
        if date_key not in sessions_by_date:
            sessions_by_date[date_key] = 0
        sessions_by_date[date_key] += 1

    # Return maximum sessions in any single day
    return max(sessions_by_date.values()) if sessions_by_date else 1


def generate_month_calendar(month: str) -> List[Dict]:
    """
    Generate a list of all days in a month with their properties.

    Args:
        month: Month in YYYY-MM format

    Returns:
        List of dictionaries with date information
    """
    month_start = datetime.strptime(month + "-01", "%Y-%m-%d")
    month_end = month_start.replace(
        month=month_start.month % 12 + 1, day=1
    ) - timedelta(days=1)

    days = []
    current_date = month_start.date()

    while current_date <= month_end.date():
        days.append(
            {
                "date": current_date,
                "date_str": current_date.strftime("%Y-%m-%d"),
                "day_name": current_date.strftime("%A"),
                "is_weekend": is_weekend(current_date),
                "formatted_date": current_date.strftime("%d-%b-%y").lstrip(
                    "0"
                ),  # "1-Jul-25"
            }
        )
        current_date += timedelta(days=1)

    return days


def organize_sessions_by_date(sessions: List[Dict]) -> Dict[str, List[Dict]]:
    """
    Organize sessions by date and sort by start time.

    Args:
        sessions: List of session dictionaries

    Returns:
        Dictionary with date as key and sorted sessions as value
    """
    sessions_by_date = {}

    for session in sessions:
        date_key = session["date"]
        if date_key not in sessions_by_date:
            sessions_by_date[date_key] = []
        sessions_by_date[date_key].append(session)

    # Sort sessions by start time for each date
    for date_key in sessions_by_date:
        sessions_by_date[date_key].sort(key=lambda x: x["start_time"])

    return sessions_by_date


# =============================================================================
# EXCEL STRUCTURE FUNCTIONS
# =============================================================================


# def generate_column_headers(max_sessions: int) -> List[str]:
#     """
#     Generate column headers for the Excel file.

#     Args:
#         max_sessions: Maximum number of sessions per day

#     Returns:
#         List of column header names
#     """
#     headers = ["Date", "Day"]

#     # Add dynamic session columns with cleaner names
#     for i in range(1, max_sessions + 1):
#         headers.extend(
#             [
#                 "Off Session",
#                 "Start Time",
#                 "End Time",
#                 "Scribe",
#                 "Patient No.",
#                 "Session Total",
#             ]
#         )

#     headers.append("DAILY TOTAL HOURS")

#     return headers


# def generate_column_headers(max_sessions: int) -> List[str]:
#     """
#     Generate column headers for the Excel file.

#     Args:
#         max_sessions: Maximum number of sessions per day (should be 2)

#     Returns:
#         List of column header names
#     """
#     headers = ["Date", "Day"]

#     # Add dynamic session columns with cleaner names
#     for i in range(1, max_sessions + 1):
#         headers.extend(
#             [
#                 "Off Session",
#                 "Start Time",
#                 "End Time",
#                 "Scribe",
#                 "Patient No.",
#                 "Session Total",
#             ]
#         )

#     # Add daily total and combined columns
#     headers.extend(
#         [
#             "DAILY TOTAL HOURS",
#             "Scribes",  # NEW: Combined scribes column
#             "Patient Nos",  # NEW: Combined patient numbers column
#         ]
#     )

#     return headers


def generate_column_headers(max_sessions: int) -> List[str]:
    """
    Generate column headers for the Excel file.

    Args:
        max_sessions: Maximum number of sessions per day (should be 2)

    Returns:
        List of column header names
    """
    headers = ["Date", "Day"]

    # Add dynamic session columns - REMOVED Scribe and Patient No. from individual sessions
    for i in range(1, max_sessions + 1):
        headers.extend(
            [
                "Off Session",
                "Start Time",
                "End Time",
                "Session Total",
            ]
        )

    # Add daily total and combined columns
    headers.extend(
        [
            "DAILY TOTAL HOURS",
            "Scribes",  # NEW: Combined scribes column
            "Patient Nos",  # NEW: Combined patient numbers column
        ]
    )

    return headers


# def generate_excel_data(
#     doctor_sessions: List[Dict], month_days: List[Dict], max_sessions: int
# ) -> List[List]:
#     """
#     Generate the data rows for the Excel file.

#     Args:
#         doctor_sessions: List of sessions for the doctor
#         month_days: List of all days in the month
#         max_sessions: Maximum number of sessions per day

#     Returns:
#         List of rows (each row is a list of cell values)
#     """
#     # Organize sessions by date
#     sessions_by_date = organize_sessions_by_date(doctor_sessions)

#     data_rows = []

#     for day_info in month_days:
#         date_str = day_info["date_str"]
#         row = [
#             day_info["formatted_date"],  # Date column
#             day_info["day_name"],  # Day column
#         ]

#         # Get sessions for this date
#         day_sessions = sessions_by_date.get(date_str, [])
#         daily_total_minutes = 0

#         # Fill session columns
#         for session_num in range(max_sessions):
#             if session_num < len(day_sessions):
#                 # There is a session for this slot
#                 session = day_sessions[session_num]

#                 # Calculate session duration
#                 try:
#                     start_time = datetime.strptime(session["start_time"], "%H:%M")
#                     end_time = datetime.strptime(session["end_time"], "%H:%M")
#                     duration_minutes = (end_time - start_time).total_seconds() / 60
#                     duration_hours = duration_minutes / 60
#                     daily_total_minutes += duration_minutes
#                 except:
#                     duration_hours = 0

#                 row.extend(
#                     [
#                         False,  # OFF_SESSION = False (there is a session)
#                         session["start_time"],
#                         session["end_time"],
#                         session["scribe_name"],
#                         f"#{session['patient_number']}",
#                         f"{duration_hours:.1f}h",
#                     ]
#                 )
#             else:
#                 # No session for this slot
#                 row.extend(
#                     [
#                         True,  # OFF_SESSION = True (no session)
#                         "",  # Start
#                         "",  # End
#                         "",  # Scribe
#                         "",  # Patient No
#                         "",  # Session Total
#                     ]
#                 )

#         # Add daily total hours
#         daily_total_hours = daily_total_minutes / 60
#         row.append(f"{daily_total_hours:.1f}h" if daily_total_hours > 0 else "0.0h")

#         data_rows.append(row)

#     return data_rows


# def generate_excel_data(
#     doctor_sessions: List[Dict], month_days: List[Dict], max_sessions: int
# ) -> List[List]:
#     """
#     Generate the data rows for the Excel file with time-based session placement.

#     Args:
#         doctor_sessions: List of sessions for the doctor
#         month_days: List of all days in the month
#         max_sessions: Maximum number of sessions per day (should be 2)

#     Returns:
#         List of rows (each row is a list of cell values)
#     """
#     # Organize sessions by date
#     sessions_by_date = organize_sessions_by_date(doctor_sessions)

#     data_rows = []

#     for day_info in month_days:
#         date_str = day_info["date_str"]
#         row = [
#             day_info["formatted_date"],  # Date column
#             day_info["day_name"],  # Day column
#         ]

#         # Get sessions for this date
#         day_sessions = sessions_by_date.get(date_str, [])
#         daily_total_minutes = 0

#         # Determine session placement based on count and time
#         session_1_data = None
#         session_2_data = None

#         if len(day_sessions) == 0:
#             # No sessions - both off
#             session_1_data = "OFF"
#             session_2_data = "OFF"

#         elif len(day_sessions) == 1:
#             # Single session - place based on time
#             session = day_sessions[0]
#             try:
#                 session_start_time = datetime.strptime(session["start_time"], "%H:%M")
#                 cutoff_time = datetime.strptime(MORNING_CUTOFF_TIME, "%H:%M")

#                 if session_start_time <= cutoff_time:
#                     # Morning session - put in Session 1
#                     session_1_data = session
#                     session_2_data = "OFF"
#                 else:
#                     # Afternoon session - put in Session 2
#                     session_1_data = "OFF"
#                     session_2_data = session
#             except ValueError:
#                 # If time parsing fails, default to Session 1
#                 session_1_data = session
#                 session_2_data = "OFF"

#         elif len(day_sessions) == 2:
#             # Two sessions - chronological order (existing behavior)
#             session_1_data = day_sessions[0]  # earliest (already sorted)
#             session_2_data = day_sessions[1]  # latest

#         else:
#             # More than 2 sessions (shouldn't happen with new limit, but handle gracefully)
#             session_1_data = day_sessions[0]
#             session_2_data = day_sessions[1]

#         # Build the row with Session 1 and Session 2 data
#         sessions_data = [session_1_data, session_2_data]

#         for session_data in sessions_data:
#             if session_data == "OFF":
#                 # OFF session
#                 row.extend(
#                     [
#                         True,  # OFF_SESSION = True
#                         OFF_SESSION_PLACEHOLDER,  # Start Time
#                         OFF_SESSION_PLACEHOLDER,  # End Time
#                         OFF_SESSION_PLACEHOLDER,  # Scribe
#                         OFF_SESSION_PLACEHOLDER,  # Patient No
#                         OFF_SESSION_PLACEHOLDER,  # Session Total
#                     ]
#                 )
#             else:
#                 # Active session
#                 session = session_data

#                 # Calculate session duration
#                 try:
#                     start_time = datetime.strptime(session["start_time"], "%H:%M")
#                     end_time = datetime.strptime(session["end_time"], "%H:%M")
#                     duration_minutes = (end_time - start_time).total_seconds() / 60
#                     duration_hours = duration_minutes / 60
#                     daily_total_minutes += duration_minutes
#                 except:
#                     duration_hours = 0

#                 row.extend(
#                     [
#                         False,  # OFF_SESSION = False (there is a session)
#                         session["start_time"],
#                         session["end_time"],
#                         session["scribe_name"],
#                         f"#{session['patient_number']}",
#                         f"{duration_hours:.1f}h",
#                     ]
#                 )

#         # Add daily total hours
#         daily_total_hours = daily_total_minutes / 60
#         row.append(f"{daily_total_hours:.1f}h" if daily_total_hours > 0 else "0.0h")

#         data_rows.append(row)

#     return data_rows


# def generate_excel_data(
#     doctor_sessions: List[Dict], month_days: List[Dict], max_sessions: int
# ) -> List[List]:
#     """
#     Generate the data rows for the Excel file with time-based session placement.

#     Args:
#         doctor_sessions: List of sessions for the doctor
#         month_days: List of all days in the month
#         max_sessions: Maximum number of sessions per day (should be 2)

#     Returns:
#         List of rows (each row is a list of cell values)
#     """
#     # Organize sessions by date
#     sessions_by_date = organize_sessions_by_date(doctor_sessions)

#     data_rows = []

#     for day_info in month_days:
#         date_str = day_info["date_str"]
#         row = [
#             day_info["formatted_date"],  # Date column
#             day_info["day_name"],  # Day column
#         ]

#         # Get sessions for this date
#         day_sessions = sessions_by_date.get(date_str, [])
#         daily_total_minutes = 0

#         # Determine session placement based on count and time
#         session_1_data = None
#         session_2_data = None

#         if len(day_sessions) == 0:
#             # No sessions - both off
#             session_1_data = "OFF"
#             session_2_data = "OFF"

#         elif len(day_sessions) == 1:
#             # Single session - place based on time
#             session = day_sessions[0]
#             try:
#                 session_start_time = datetime.strptime(session["start_time"], "%H:%M")
#                 cutoff_time = datetime.strptime(MORNING_CUTOFF_TIME, "%H:%M")

#                 if session_start_time <= cutoff_time:
#                     # Morning session - put in Session 1
#                     session_1_data = session
#                     session_2_data = "OFF"
#                 else:
#                     # Afternoon session - put in Session 2
#                     session_1_data = "OFF"
#                     session_2_data = session
#             except ValueError:
#                 # If time parsing fails, default to Session 1
#                 session_1_data = session
#                 session_2_data = "OFF"

#         elif len(day_sessions) == 2:
#             # Two sessions - chronological order (existing behavior)
#             session_1_data = day_sessions[0]  # earliest (already sorted)
#             session_2_data = day_sessions[1]  # latest

#         else:
#             # More than 2 sessions (shouldn't happen with new limit, but handle gracefully)
#             session_1_data = day_sessions[0]
#             session_2_data = day_sessions[1]

#         # Build the row with Session 1 and Session 2 data
#         sessions_data = [session_1_data, session_2_data]

#         # Track scribe names and patient numbers for combined columns
#         active_scribes = []
#         active_patient_nos = []

#         for session_data in sessions_data:
#             if session_data == "OFF":
#                 # OFF session
#                 row.extend(
#                     [
#                         True,  # OFF_SESSION = True
#                         OFF_SESSION_PLACEHOLDER,  # Start Time
#                         OFF_SESSION_PLACEHOLDER,  # End Time
#                         OFF_SESSION_PLACEHOLDER,  # Scribe
#                         OFF_SESSION_PLACEHOLDER,  # Patient No
#                         OFF_SESSION_PLACEHOLDER,  # Session Total
#                     ]
#                 )
#                 # Don't add to combined columns (skip the "-")
#             else:
#                 # Active session
#                 session = session_data

#                 # Calculate session duration
#                 try:
#                     start_time = datetime.strptime(session["start_time"], "%H:%M")
#                     end_time = datetime.strptime(session["end_time"], "%H:%M")
#                     duration_minutes = (end_time - start_time).total_seconds() / 60
#                     duration_hours = duration_minutes / 60
#                     daily_total_minutes += duration_minutes
#                 except:
#                     duration_hours = 0

#                 # Add to individual session columns
#                 patient_no_formatted = f"#{session['patient_number']}"
#                 row.extend(
#                     [
#                         False,  # OFF_SESSION = False (there is a session)
#                         session["start_time"],
#                         session["end_time"],
#                         session["scribe_name"],
#                         patient_no_formatted,
#                         f"{duration_hours:.1f}h",
#                     ]
#                 )

#                 # Add to combined columns data
#                 active_scribes.append(session["scribe_name"])
#                 active_patient_nos.append(patient_no_formatted)

#         # Add daily total hours
#         daily_total_hours = daily_total_minutes / 60
#         row.append(f"{daily_total_hours:.1f}h" if daily_total_hours > 0 else "0.0h")

#         # Add combined columns
#         if active_scribes:
#             combined_scribes = ", ".join(active_scribes)
#             combined_patient_nos = ", ".join(active_patient_nos)
#         else:
#             # No active sessions
#             combined_scribes = OFF_SESSION_PLACEHOLDER
#             combined_patient_nos = OFF_SESSION_PLACEHOLDER

#         row.extend(
#             [
#                 combined_scribes,  # Combined Scribes column
#                 combined_patient_nos,  # Combined Patient Numbers column
#             ]
#         )

#         data_rows.append(row)

#     return data_rows


def generate_excel_data(
    doctor_sessions: List[Dict], month_days: List[Dict], max_sessions: int
) -> List[List]:
    """
    Generate the data rows for the Excel file with time-based session placement.

    Args:
        doctor_sessions: List of sessions for the doctor
        month_days: List of all days in the month
        max_sessions: Maximum number of sessions per day (should be 2)

    Returns:
        List of rows (each row is a list of cell values)
    """
    # Organize sessions by date
    sessions_by_date = organize_sessions_by_date(doctor_sessions)

    data_rows = []

    for day_info in month_days:
        date_str = day_info["date_str"]
        row = [
            day_info["formatted_date"],  # Date column
            day_info["day_name"],  # Day column
        ]

        # Get sessions for this date
        day_sessions = sessions_by_date.get(date_str, [])
        daily_total_minutes = 0

        # Determine session placement based on count and time
        session_1_data = None
        session_2_data = None

        if len(day_sessions) == 0:
            # No sessions - both off
            session_1_data = "OFF"
            session_2_data = "OFF"

        elif len(day_sessions) == 1:
            # Single session - place based on time
            session = day_sessions[0]
            try:
                session_start_time = datetime.strptime(session["start_time"], "%H:%M")
                cutoff_time = datetime.strptime(MORNING_CUTOFF_TIME, "%H:%M")

                if session_start_time <= cutoff_time:
                    # Morning session - put in Session 1
                    session_1_data = session
                    session_2_data = "OFF"
                else:
                    # Afternoon session - put in Session 2
                    session_1_data = "OFF"
                    session_2_data = session
            except ValueError:
                # If time parsing fails, default to Session 1
                session_1_data = session
                session_2_data = "OFF"

        elif len(day_sessions) == 2:
            # Two sessions - chronological order (existing behavior)
            session_1_data = day_sessions[0]  # earliest (already sorted)
            session_2_data = day_sessions[1]  # latest

        else:
            # More than 2 sessions (shouldn't happen with new limit, but handle gracefully)
            session_1_data = day_sessions[0]
            session_2_data = day_sessions[1]

        # Build the row with Session 1 and Session 2 data
        sessions_data = [session_1_data, session_2_data]

        # Track scribe names and patient numbers for combined columns
        active_scribes = []
        active_patient_nos = []

        for session_data in sessions_data:
            if session_data == "OFF":
                # OFF session - REMOVED Scribe and Patient No. columns
                row.extend(
                    [
                        True,  # OFF_SESSION = True
                        OFF_SESSION_PLACEHOLDER,  # Start Time
                        OFF_SESSION_PLACEHOLDER,  # End Time
                        OFF_SESSION_PLACEHOLDER,  # Session Total
                    ]
                )
                # Don't add to combined columns (skip the "-")
            else:
                # Active session
                session = session_data

                # Calculate session duration
                try:
                    start_time = datetime.strptime(session["start_time"], "%H:%M")
                    end_time = datetime.strptime(session["end_time"], "%H:%M")
                    duration_minutes = (end_time - start_time).total_seconds() / 60
                    duration_hours = duration_minutes / 60
                    daily_total_minutes += duration_minutes
                except:
                    duration_hours = 0

                # Add to individual session columns - REMOVED Scribe and Patient No.
                patient_no_formatted = f"#{session['patient_number']}"
                row.extend(
                    [
                        False,  # OFF_SESSION = False (there is a session)
                        session["start_time"],
                        session["end_time"],
                        f"{duration_hours:.1f}h",
                    ]
                )

                # Add to combined columns data
                active_scribes.append(session["scribe_name"])
                active_patient_nos.append(patient_no_formatted)

        # Add daily total hours
        daily_total_hours = daily_total_minutes / 60
        row.append(f"{daily_total_hours:.1f}h" if daily_total_hours > 0 else "0.0h")

        # Add combined columns
        if active_scribes:
            combined_scribes = ", ".join(active_scribes)
            combined_patient_nos = ", ".join(active_patient_nos)
        else:
            # No active sessions
            combined_scribes = OFF_SESSION_PLACEHOLDER
            combined_patient_nos = OFF_SESSION_PLACEHOLDER

        row.extend(
            [
                combined_scribes,  # Combined Scribes column
                combined_patient_nos,  # Combined Patient Numbers column
            ]
        )

        data_rows.append(row)

    return data_rows


def calculate_monthly_total(doctor_sessions: List[Dict]) -> float:
    """
    Calculate total monthly hours for a doctor.

    Args:
        doctor_sessions: List of sessions for the doctor

    Returns:
        Total hours for the month
    """
    duration_info = calculate_total_duration(doctor_sessions)
    return duration_info["total_hours"]


# =============================================================================
# EXCEL FORMATTING FUNCTIONS
# =============================================================================


def create_excel_workbook(
    team_leader_name: str, month: str, doctor_data: Dict[str, Any]
) -> Workbook:
    """
    Create and format the Excel workbook with multiple doctor sheets.

    Args:
        team_leader_name: Name of the team leader
        month: Month in YYYY-MM format
        doctor_data: Dictionary with doctor info and their session data

    Returns:
        Formatted Excel workbook
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create a sheet for each doctor
    for doctor_id, data in doctor_data.items():
        doctor_info = data["doctor_info"]
        sessions = data["sessions"]
        max_sessions = data["max_sessions"]
        month_days = data["month_days"]

        # Create sheet with doctor name
        sheet_name = format_doctor_first_last_name(doctor_info)[
            :31
        ]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)

        # Format the sheet
        format_doctor_sheet(ws, doctor_info, month, sessions, max_sessions, month_days)

    return wb


def format_doctor_sheet(
    ws,
    doctor_info: Dict,
    month: str,
    sessions: List[Dict],
    max_sessions: int,
    month_days: List[Dict],
):
    """
    Format a single doctor sheet with headers, data, and styling.

    Args:
        ws: Worksheet object
        doctor_info: Doctor information dictionary
        month: Month in YYYY-MM format
        sessions: List of sessions for this doctor
        max_sessions: Maximum sessions per day
        month_days: List of all days in the month
    """
    logger.info(
        f"Formatting sheet for doctor: {format_doctor_first_last_name(doctor_info)}"
    )
    logger.info(f"Sessions: {len(sessions)}, Max sessions/day: {max_sessions}")

    # Color definitions - EXACT COLORS FROM SCREENSHOT
    weekend_fill = PatternFill(
        start_color="87CEEB", end_color="87CEEB", fill_type="solid"
    )  # Sky blue for weekends

    total_monthly_hours_fill = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )  # Light green for monthly total

    session_header_fill = PatternFill(
        start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
    )  # Lavender for session headers

    # Font definitions
    header_font = Font(bold=True, size=11)
    total_monthly_hours_font = Font(bold=True, size=14)
    month_display_font = Font(bold=True, size=14)
    doctor_name_font = Font(bold=True, size=14)

    # Border definition - thin borders on all sides
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Month and year for display
    month_obj = datetime.strptime(month + "-01", "%Y-%m-%d")
    month_display = month_obj.strftime("%B %Y")

    # Calculate monthly total
    monthly_total = calculate_monthly_total(sessions)
    logger.info(f"Monthly total hours: {monthly_total}")

    try:
        # DATE - merged vertically A1:B2
        ws.merge_cells("A1:B2")
        month_display_cell = ws["A1"]
        month_display_cell.value = month_display
        month_display_cell.font = month_display_font
        month_display_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Apply borders to all cells in merged range A1:B2
        for row in range(1, 3):  # Row 1 to Row 2
            for col in range(1, 3):  # Column A to Column B
                ws.cell(row=row, column=col).border = thin_border

        # DOCTOR NAME - merged horizontally C1:D1
        ws.merge_cells("C1:D1")
        doctor_name_index_cell = ws["C1"]
        doctor_name_index_cell.value = "DOCTOR NAME:"
        doctor_name_index_cell.font = doctor_name_font
        # Apply borders to all cells in merged range C1:D1
        for col in range(3, 5):  # C1 to D1
            ws.cell(row=1, column=col).border = thin_border

        # Doctor's full name in E1
        ws.merge_cells("E1:F1")
        doctor_name_cell = ws["E1"]
        doctor_name_cell.value = format_doctor_first_last_name(doctor_info)
        doctor_name_cell.font = doctor_name_font
        doctor_name_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Apply borders to all cells in merged range E1:F1
        for col in range(5, 7):  # E1 to F1
            ws.cell(row=1, column=col).border = thin_border

        # Total Monthly Hours - merged horizontally G1:H1
        ws.merge_cells("G1:H1")
        total_monthly_cell = ws["G1"]
        total_monthly_cell.value = "Total Monthly Hours"
        total_monthly_cell.font = total_monthly_hours_font
        total_monthly_cell.fill = total_monthly_hours_fill
        total_monthly_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Apply borders and fill to merged range G1:H1
        for col in range(7, 9):  # G1 to H1
            cell = ws.cell(row=1, column=col)
            cell.border = thin_border
            cell.fill = total_monthly_hours_fill

        # Monthly total value in I1
        ws["I1"] = f"{monthly_total:.1f}h"
        ws["I1"].font = total_monthly_hours_font
        ws["I1"].border = thin_border
        ws["I1"].fill = total_monthly_hours_fill

        # Apply borders to remaining cells in row 1
        max_col = len(generate_column_headers(max_sessions))
        for col in range(10, max_col + 1):
            ws.cell(row=1, column=col).border = thin_border

        # Apply borders to row 2 cells that aren't part of the date merge
        for col in range(3, max_col + 1):
            ws.cell(row=2, column=col).border = thin_border

        # Row 2: Session headers
        session_headers_row = 2
        col_start = 3  # Column C

        # First, apply borders to columns A and B in row 2
        ws.cell(row=2, column=1).border = thin_border
        ws.cell(row=2, column=2).border = thin_border

        for i in range(max_sessions):
            session_start_col = col_start + (i * 6)
            session_end_col = session_start_col + 5

            # Merge cells for session header
            ws.merge_cells(
                start_row=session_headers_row,
                start_column=session_start_col,
                end_row=session_headers_row,
                end_column=session_end_col,
            )

            cell = ws.cell(row=session_headers_row, column=session_start_col)
            cell.value = f"SESSION {i + 1}"
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.fill = session_header_fill
            cell.border = thin_border

            # Apply borders and fill to all cells in the merged range
            for col in range(session_start_col, session_end_col + 1):
                range_cell = ws.cell(row=session_headers_row, column=col)
                range_cell.border = thin_border
                range_cell.fill = session_header_fill

        # Apply borders to remaining cells in row 2 (after sessions)
        last_session_col = col_start + (max_sessions * 6) - 1
        max_col = len(generate_column_headers(max_sessions))
        for col in range(last_session_col + 1, max_col + 1):
            ws.cell(row=2, column=col).border = thin_border

        # Row 3: Column headers
        headers = generate_column_headers(max_sessions)
        logger.info(f"Generated {len(headers)} column headers")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal="center")
            cell.font = header_font
            cell.border = thin_border

        # Data rows (starting from row 5)
        data_rows = generate_excel_data(sessions, month_days, max_sessions)
        logger.info(f"Generated {len(data_rows)} data rows")

        for row_num, row_data in enumerate(data_rows, 4):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = thin_border  # Add border to every data cell

                # Apply weekend highlighting
                day_info = month_days[row_num - 5]  # Adjust for 0-based index
                if day_info["is_weekend"]:
                    cell.fill = weekend_fill

        # Auto-adjust column widths
        logger.info("Auto-adjusting column widths...")

        # Get the actual number of columns with data
        max_col = len(headers)

        for col_num in range(1, max_col + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)

            # Check all rows for this column to find max length
            for row_num in range(
                1, len(data_rows) + 4
            ):  # +4 for header rows (now 3 header rows + 1)
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception as e:
                    logger.debug(f"Error checking cell {column_letter}{row_num}: {e}")
                    continue

            # Set column width
            adjusted_width = min(max(max_length + 2, 8), 25)  # Min 8, Max 25 characters
            ws.column_dimensions[column_letter].width = adjusted_width
            logger.debug(f"Column {column_letter}: width set to {adjusted_width}")

        logger.info("Sheet formatting completed successfully")

    except Exception as e:
        logger.error(f"Error formatting doctor sheet: {e}")
        raise


# =============================================================================
# MAIN EXPORT FUNCTIONS
# =============================================================================


def prepare_export_data(team_leader_id: str, month: str) -> Dict[str, Any]:
    """
    Prepare all data needed for export.

    Args:
        team_leader_id: Team leader ID
        month: Month in YYYY-MM format

    Returns:
        Dictionary with organized export data
    """
    # Get all month data
    doctor_sessions = get_month_data(team_leader_id, month)

    # Generate month calendar
    month_days = generate_month_calendar(month)

    # Prepare data for each doctor
    export_data = {}

    for doctor in POC_DOCTORS:
        doctor_id = doctor["id"]
        sessions = doctor_sessions.get(doctor_id, [])
        max_sessions = calculate_max_sessions_per_day(sessions)

        export_data[doctor_id] = {
            "doctor_info": doctor,
            "sessions": sessions,
            "max_sessions": max_sessions,
            "month_days": month_days,
        }

    return export_data


def generate_excel_file(
    team_leader_name: str, month: str, export_data: Dict[str, Any]
) -> io.BytesIO:
    """
    Generate the complete Excel file.

    Args:
        team_leader_name: Name of the team leader
        month: Month in YYYY-MM format
        export_data: Prepared export data

    Returns:
        BytesIO object containing the Excel file
    """
    logger.info(f"Starting Excel file generation for {team_leader_name}, {month}")
    logger.info(f"Export data contains {len(export_data)} doctors")

    try:
        # Create workbook
        wb = create_excel_workbook(team_leader_name, month, export_data)
        logger.info("Workbook created successfully")

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        logger.info("Excel file generated successfully")
        return excel_file

    except Exception as e:
        logger.error(f"Error generating Excel file: {e}")
        raise


# =============================================================================
# STREAMLIT UI FUNCTIONS
# =============================================================================


def render_export_filters() -> Tuple[str, str]:
    """
    Render the filter section for export parameters.

    Returns:
        Tuple of (selected_team_leader, selected_month)
    """
    st.subheader("📋 Export Parameters")

    with st.container(border=True):
        col1, col2 = st.columns(2)

        with col1:
            # Team Leader selection
            tl_options = [POC_TEAM_LEADER["name"]]  # For now, only one TL
            selected_tl = st.selectbox(
                "Team Leader:", options=tl_options, key="export_tl_filter"
            )

        with col2:
            # Month selection
            current_date = datetime.now()
            month_options = []

            # Generate last 12 months + next 6 months
            for i in range(-12, 7):
                month_date = current_date.replace(day=1) + timedelta(days=32 * i)
                month_date = month_date.replace(day=1)
                month_label = month_date.strftime("%B %Y")
                month_value = month_date.strftime("%Y-%m")
                month_options.append((month_label, month_value))

            # Default to current month
            current_month_value = current_date.strftime("%Y-%m")

            month_labels = [opt[0] for opt in month_options]
            month_values = [opt[1] for opt in month_options]

            try:
                default_index = month_values.index(current_month_value)
            except ValueError:
                default_index = 0

            selected_month_label = st.selectbox(
                "Month:",
                options=month_labels,
                index=default_index,
                key="export_month_filter",
            )

            selected_month = month_values[month_labels.index(selected_month_label)]

    return selected_tl, selected_month


def render_export_preview(export_data: Dict[str, Any], month: str):
    """
    Render a preview of what will be exported.

    Args:
        export_data: Prepared export data
        month: Month in YYYY-MM format
    """
    st.subheader("📊 Export Preview")

    with st.container(border=True):
        if not export_data:
            st.info("No data available for the selected month.")
            return

        # Summary stats
        total_doctors = len(export_data)
        total_sessions = sum(len(data["sessions"]) for data in export_data.values())
        max_sessions_overall = (
            max(data["max_sessions"] for data in export_data.values())
            if export_data
            else 0
        )

        col1, col2, col3 = st.columns(3)

        with col1:
            st.metric("Doctors", total_doctors)

        with col2:
            st.metric("Total Sessions", total_sessions)

        with col3:
            st.metric("Max Sessions/Day", max_sessions_overall)

        # Doctor breakdown
        st.write("**Doctor Sessions Breakdown:**")
        for doctor_id, data in export_data.items():
            doctor_name = format_doctor_name(data["doctor_info"])
            session_count = len(data["sessions"])
            max_sessions = data["max_sessions"]

            if session_count > 0:
                monthly_total = calculate_monthly_total(data["sessions"])
                st.write(
                    f"• **{doctor_name}**: {session_count} sessions, {monthly_total:.1f}h total, max {max_sessions} sessions/day"
                )
            else:
                st.write(f"• **{doctor_name}**: No sessions scheduled")


def render_export_section(
    team_leader_name: str, month: str, export_data: Dict[str, Any]
):
    """
    Render the export button and handle file generation.

    Args:
        team_leader_name: Name of the team leader
        month: Month in YYYY-MM format
        export_data: Prepared export data
    """
    st.subheader("📤 Generate Export")

    with st.container(border=True):
        if not export_data or not any(
            data["sessions"] for data in export_data.values()
        ):
            st.warning("No session data available for export.")
            st.button("📊 Generate Excel File", disabled=True, help="No data to export")
            return

        # File name preview
        month_obj = datetime.strptime(month + "-01", "%Y-%m-%d")
        file_name = (
            f"{month_obj.strftime('%B_%Y')}_{team_leader_name.replace(' ', '_')}.xlsx"
        )
        st.write(f"**File name:** `{file_name}`")

        # Export button
        if st.button(
            "📊 Generate Excel File", type="primary", use_container_width=True
        ):
            with st.spinner("Generating Excel file..."):
                try:
                    excel_file = generate_excel_file(
                        team_leader_name, month, export_data
                    )

                    st.success("✅ Excel file generated successfully!")

                    # Download button
                    st.download_button(
                        label="⬇️ Download Excel File",
                        data=excel_file,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )

                except Exception as e:
                    st.error(f"❌ Error generating Excel file: {str(e)}")


# =============================================================================
# MAIN EXPORT PAGE
# =============================================================================


def render_export_page():
    """Main export page function"""
    st.title("📤 Excel Export")

    # Filters section
    selected_tl, selected_month = render_export_filters()

    st.markdown("---")

    # Prepare export data
    with st.spinner("Loading data..."):
        export_data = prepare_export_data(POC_TEAM_LEADER["username"], selected_month)

    # Preview section
    render_export_preview(export_data, selected_month)

    st.markdown("---")

    # Export section
    render_export_section(selected_tl, selected_month, export_data)


if __name__ == "__main__":
    render_export_page()
