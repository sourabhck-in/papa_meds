# src/ui/pages/export.py
"""
Excel Export Functionality for Medical Schedule Management System

Generate Excel files matching the template format with dynamic session columns
with proper logging and modular architecture.
"""

import sys
import os
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import calendar
from typing import Dict, List, Any, Tuple, Optional
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Import our modular components
try:
    from src.utils.logging_config import get_logger, setup_logging
    from src.utils.config import (
        APP_ICON,
        MORNING_CUTOFF_TIME,
        OFF_SESSION_PLACEHOLDER,
        POC_TEAM_LEADER,
        POC_DOCTORS,
        WEEKEND_DAYS,
        get_streamlit_config,
        is_debug_mode,
    )
    from src.infrastructure.sheets.sheets_ops import (
        load_submissions,
        format_doctor_name,
        format_doctor_first_last_name,
        validate_sheets_connection,
        get_sheets_stats,
    )
    from src.core.validators.validation import (
        is_weekend,
        calculate_total_duration,
    )
except ImportError as e:
    print(f"Import error in export.py: {e}")
    st.error(f"Import error: {e}")
    st.stop()

# Initialize logging
setup_logging()
logger = get_logger(__name__)


# =============================================================================
# APP CONFIGURATION
# =============================================================================


def setup_export_app() -> None:
    """Configure Streamlit app for export page"""
    logger.info("Setting up export page...")

    try:
        config = get_streamlit_config()
        config["page_title"] = "Excel Export"
        st.set_page_config(**config)

        # Validate Google Sheets connection
        if not validate_sheets_connection():
            st.error("❌ Google Sheets connection failed. Cannot load export data.")
            logger.error("Google Sheets connection failed on export page")
            st.stop()

        logger.info("Export page setup complete")

    except Exception as e:
        logger.error(f"Failed to setup export page: {str(e)}")
        st.error(f"❌ Export page setup failed: {str(e)}")
        st.stop()


def render_export_header() -> None:
    """Render export page header"""
    logger.debug("Rendering export page header")

    st.title("📤 Excel Export")

    if is_debug_mode():
        with st.expander("🔧 Debug Information", expanded=False):
            render_export_debug_info()


def render_export_debug_info() -> None:
    """Render debug information for export page"""
    try:
        stats = get_sheets_stats()
        st.json({"sheets_stats": stats})
    except Exception as e:
        st.error(f"Debug info error: {str(e)}")
        logger.error(f"Error rendering export debug info: {str(e)}")


# =============================================================================
# DATA PROCESSING FUNCTIONS
# =============================================================================


def get_month_export_data(team_leader_id: str, month: str) -> Dict[str, List[Dict]]:
    """
    Get all submissions data for a specific team leader and month.

    Args:
        team_leader_id: Team leader ID
        month: Month in YYYY-MM format

    Returns:
        Dictionary with doctor_id as keys and list of sessions as values
    """
    logger.info(f"Loading month export data for TL: {team_leader_id}, Month: {month}")

    try:
        # Load all submissions from Google Sheets
        all_submissions = load_submissions()
        logger.info(
            f"Loaded {len(all_submissions)} total submissions from Google Sheets"
        )

        if not all_submissions:
            logger.warning("No submissions found for export")
            return {}

        # Filter by month
        month_start = datetime.strptime(month + "-01", "%Y-%m-%d")
        month_end = month_start.replace(
            month=month_start.month % 12 + 1, day=1
        ) - timedelta(days=1)

        logger.info(
            f"Filtering for date range: {month_start.date()} to {month_end.date()}"
        )

        # Group sessions by doctor
        doctor_sessions = {}
        filtered_count = 0

        for submission in all_submissions:
            try:
                submission_date = datetime.strptime(submission["date"], "%Y-%m-%d")

                # Check if submission is in the target month
                if month_start.date() <= submission_date.date() <= month_end.date():
                    doctor_id = submission["doctor_id"]

                    if doctor_id not in doctor_sessions:
                        doctor_sessions[doctor_id] = []

                    doctor_sessions[doctor_id].append(submission)
                    filtered_count += 1

            except ValueError as e:
                logger.error(
                    f"Invalid date format in submission: {submission.get('date')} - {e}"
                )
                continue

        logger.info(f"Filtered to {filtered_count} submissions for export")
        logger.info(
            f"Sessions by doctor: {[(doc_id, len(sessions)) for doc_id, sessions in doctor_sessions.items()]}"
        )

        return doctor_sessions

    except Exception as e:
        logger.error(f"Error getting month export data: {str(e)}")
        return {}


def calculate_max_sessions_per_day(sessions: List[Dict]) -> int:
    """
    Calculate the maximum number of sessions per day for a given set of sessions.

    Args:
        sessions: List of session dictionaries

    Returns:
        Maximum number of sessions in a single day
    """
    logger.debug(f"Calculating max sessions per day for {len(sessions)} sessions")

    if not sessions:
        return 1  # Minimum 1 session column

    # Group sessions by date
    sessions_by_date = {}
    for session in sessions:
        date_key = session["date"]
        sessions_by_date[date_key] = sessions_by_date.get(date_key, 0) + 1

    max_sessions = max(sessions_by_date.values()) if sessions_by_date else 1
    logger.debug(f"Max sessions per day: {max_sessions}")
    return max_sessions


def generate_month_calendar(month: str) -> List[Dict]:
    """
    Generate a list of all days in a month with their properties.

    Args:
        month: Month in YYYY-MM format

    Returns:
        List of dictionaries with date information
    """
    logger.debug(f"Generating month calendar for {month}")

    month_start = datetime.strptime(month + "-01", "%Y-%m-%d")
    month_end = month_start.replace(
        month=month_start.month % 12 + 1, day=1
    ) - timedelta(days=1)

    days = []
    current_date = month_start.date()

    while current_date <= month_end.date():
        days.append(
            {
                "date": current_date,
                "date_str": current_date.strftime("%Y-%m-%d"),
                "day_name": current_date.strftime("%A"),
                "is_weekend": is_weekend(current_date),
                "formatted_date": current_date.strftime("%d-%b-%y").lstrip(
                    "0"
                ),  # "1-Jul-25"
            }
        )
        current_date += timedelta(days=1)

    logger.debug(f"Generated calendar with {len(days)} days")
    return days


def organize_sessions_by_date(sessions: List[Dict]) -> Dict[str, List[Dict]]:
    """
    Organize sessions by date and sort by start time.

    Args:
        sessions: List of session dictionaries

    Returns:
        Dictionary with date as key and sorted sessions as value
    """
    logger.debug(f"Organizing {len(sessions)} sessions by date")

    sessions_by_date = {}

    for session in sessions:
        date_key = session["date"]
        if date_key not in sessions_by_date:
            sessions_by_date[date_key] = []
        sessions_by_date[date_key].append(session)

    # Sort sessions by start time for each date
    for date_key in sessions_by_date:
        sessions_by_date[date_key].sort(key=lambda x: x["start_time"])

    logger.debug(f"Organized sessions into {len(sessions_by_date)} dates")
    return sessions_by_date


# =============================================================================
# EXCEL GENERATION FUNCTIONS
# =============================================================================


def generate_excel_column_headers(max_sessions: int) -> List[str]:
    """
    Generate column headers for the Excel file.

    Args:
        max_sessions: Maximum number of sessions per day

    Returns:
        List of column header names
    """
    logger.debug(f"Generating Excel column headers for {max_sessions} max sessions")

    headers = ["Date", "Day"]

    # Add dynamic session columns
    for i in range(1, max_sessions + 1):
        headers.extend(
            [
                "Off Session",
                "Start Time",
                "End Time",
                "Session Total",
            ]
        )

    # Add daily total and combined columns
    headers.extend(
        [
            "DAILY TOTAL HOURS",
            "Scribes",  # Combined scribes column
            "Patient Nos",  # Combined patient numbers column
        ]
    )

    logger.debug(f"Generated {len(headers)} column headers")
    return headers


def determine_session_placement(day_sessions: List[Dict], max_sessions: int) -> List:
    """
    Determine how to place sessions in the Excel columns based on time and count.

    Args:
        day_sessions: List of sessions for the day
        max_sessions: Maximum sessions per day

    Returns:
        List of session data or "OFF" strings
    """
    session_data_list = []

    if len(day_sessions) == 0:
        # No sessions - all off
        session_data_list = ["OFF"] * max_sessions

    elif len(day_sessions) == 1:
        # Single session - place based on time
        session = day_sessions[0]
        try:
            session_start_time = datetime.strptime(session["start_time"], "%H:%M")
            cutoff_time = datetime.strptime(MORNING_CUTOFF_TIME, "%H:%M")

            if session_start_time <= cutoff_time:
                # Morning session - put in Session 1
                session_data_list = [session] + ["OFF"] * (max_sessions - 1)
            else:
                # Afternoon session - put in Session 2
                session_data_list = ["OFF", session] + ["OFF"] * (max_sessions - 2)
        except ValueError:
            # If time parsing fails, default to Session 1
            session_data_list = [session] + ["OFF"] * (max_sessions - 1)

    elif len(day_sessions) >= max_sessions:
        # Multiple sessions - take first max_sessions in chronological order
        session_data_list = day_sessions[:max_sessions]

    else:
        # Fill available slots chronologically
        session_data_list = day_sessions + ["OFF"] * (max_sessions - len(day_sessions))

    return session_data_list


def calculate_session_duration_hours(session: Dict) -> float:
    """
    Calculate session duration in hours.

    Args:
        session: Session dictionary

    Returns:
        Duration in hours
    """
    try:
        start_time = datetime.strptime(session["start_time"], "%H:%M")
        end_time = datetime.strptime(session["end_time"], "%H:%M")
        duration_minutes = (end_time - start_time).total_seconds() / 60
        return duration_minutes / 60
    except Exception as e:
        logger.warning(f"Error calculating session duration: {str(e)}")
        return 0


def generate_excel_data_rows(
    doctor_sessions: List[Dict], month_days: List[Dict], max_sessions: int
) -> List[List]:
    """
    Generate the data rows for the Excel file with time-based session placement.

    Args:
        doctor_sessions: List of sessions for the doctor
        month_days: List of all days in the month
        max_sessions: Maximum number of sessions per day

    Returns:
        List of rows (each row is a list of cell values)
    """
    logger.debug(
        f"Generating Excel data rows for {len(doctor_sessions)} sessions across {len(month_days)} days"
    )

    # Organize sessions by date
    sessions_by_date = organize_sessions_by_date(doctor_sessions)
    data_rows = []

    for day_info in month_days:
        date_str = day_info["date_str"]
        row = [
            day_info["formatted_date"],  # Date column
            day_info["day_name"],  # Day column
        ]

        # Get sessions for this date
        day_sessions = sessions_by_date.get(date_str, [])
        daily_total_minutes = 0

        # Determine session placement
        session_data_list = determine_session_placement(day_sessions, max_sessions)

        # Track combined data
        active_scribes = []
        active_patient_nos = []

        # Build row data for each session slot
        for session_data in session_data_list:
            if session_data == "OFF":
                # OFF session
                row.extend(
                    [
                        True,  # OFF_SESSION = True
                        OFF_SESSION_PLACEHOLDER,  # Start Time
                        OFF_SESSION_PLACEHOLDER,  # End Time
                        OFF_SESSION_PLACEHOLDER,  # Session Total
                    ]
                )
            else:
                # Active session
                session = session_data

                # Calculate session duration
                duration_hours = calculate_session_duration_hours(session)
                daily_total_minutes += duration_hours * 60

                # Add to individual session columns
                patient_no_formatted = f"#{session['patient_number']}"
                row.extend(
                    [
                        False,  # OFF_SESSION = False
                        session["start_time"],
                        session["end_time"],
                        f"{duration_hours:.1f}h",
                    ]
                )

                # Add to combined columns data
                active_scribes.append(session["scribe_name"])
                active_patient_nos.append(patient_no_formatted)

        # Add daily total hours
        daily_total_hours = daily_total_minutes / 60
        row.append(f"{daily_total_hours:.1f}h" if daily_total_hours > 0 else "0.0h")

        # Add combined columns
        if active_scribes:
            combined_scribes = ", ".join(active_scribes)
            combined_patient_nos = ", ".join(active_patient_nos)
        else:
            combined_scribes = OFF_SESSION_PLACEHOLDER
            combined_patient_nos = OFF_SESSION_PLACEHOLDER

        row.extend([combined_scribes, combined_patient_nos])
        data_rows.append(row)

    logger.debug(f"Generated {len(data_rows)} data rows")
    return data_rows


def calculate_monthly_total_hours(doctor_sessions: List[Dict]) -> float:
    """
    Calculate total monthly hours for a doctor.

    Args:
        doctor_sessions: List of sessions for the doctor

    Returns:
        Total hours for the month
    """
    logger.debug(f"Calculating monthly total for {len(doctor_sessions)} sessions")

    duration_info = calculate_total_duration(doctor_sessions)
    total_hours = duration_info["total_hours"]

    logger.debug(f"Monthly total: {total_hours} hours")
    return total_hours


def prepare_export_data(team_leader_id: str, month: str) -> Dict[str, Any]:
    """
    Prepare all data needed for Excel export.

    Args:
        team_leader_id: Team leader ID
        month: Month in YYYY-MM format

    Returns:
        Dictionary with organized export data
    """
    logger.info(f"Preparing export data for TL: {team_leader_id}, Month: {month}")

    try:
        # Get all month data from Google Sheets
        doctor_sessions = get_month_export_data(team_leader_id, month)

        # Generate month calendar
        month_days = generate_month_calendar(month)

        # Prepare data for each doctor
        export_data = {}

        for doctor in POC_DOCTORS:
            doctor_id = doctor["id"]
            sessions = doctor_sessions.get(doctor_id, [])
            max_sessions = calculate_max_sessions_per_day(sessions)

            export_data[doctor_id] = {
                "doctor_info": doctor,
                "sessions": sessions,
                "max_sessions": max_sessions,
                "month_days": month_days,
            }

            logger.debug(
                f"Prepared data for doctor {doctor_id}: {len(sessions)} sessions"
            )

        logger.info(f"Export data preparation complete for {len(export_data)} doctors")
        return export_data

    except Exception as e:
        logger.error(f"Error preparing export data: {str(e)}")
        return {}


# =============================================================================
# EXCEL FORMATTING FUNCTIONS
# =============================================================================


def create_formatted_excel_workbook(
    team_leader_name: str, month: str, doctor_data: Dict[str, Any]
) -> Workbook:
    """
    Create and format the Excel workbook with multiple doctor sheets.

    Args:
        team_leader_name: Name of the team leader
        month: Month in YYYY-MM format
        doctor_data: Dictionary with doctor info and their session data

    Returns:
        Formatted Excel workbook
    """
    logger.info(f"Creating Excel workbook for {team_leader_name}, {month}")
    logger.info(f"Doctor data contains {len(doctor_data)} doctors")

    try:
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create a sheet for each doctor
        for doctor_id, data in doctor_data.items():
            doctor_info = data["doctor_info"]
            sessions = data["sessions"]
            max_sessions = data["max_sessions"]
            month_days = data["month_days"]

            # Create sheet with doctor name
            sheet_name = format_doctor_first_last_name(doctor_info)[
                :31
            ]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)

            logger.debug(f"Creating sheet for doctor: {sheet_name}")

            # Format the sheet
            format_doctor_excel_sheet(
                ws, doctor_info, month, sessions, max_sessions, month_days
            )

        logger.info("Excel workbook creation complete")
        return wb

    except Exception as e:
        logger.error(f"Error creating Excel workbook: {str(e)}")
        raise


def format_doctor_excel_sheet(
    ws,
    doctor_info: Dict,
    month: str,
    sessions: List[Dict],
    max_sessions: int,
    month_days: List[Dict],
) -> None:
    """
    Format a single doctor sheet with headers, data, and styling.

    Args:
        ws: Worksheet object
        doctor_info: Doctor information dictionary
        month: Month in YYYY-MM format
        sessions: List of sessions for this doctor
        max_sessions: Maximum sessions per day
        month_days: List of all days in the month
    """
    logger.debug(
        f"Formatting Excel sheet for doctor: {format_doctor_first_last_name(doctor_info)}"
    )

    try:
        # Apply Excel styling
        apply_excel_styling(ws, doctor_info, month, sessions, max_sessions)

        # Add data rows
        add_excel_data_rows(ws, sessions, month_days, max_sessions)

        # Auto-adjust column widths
        auto_adjust_column_widths(ws, max_sessions)

        logger.debug("Excel sheet formatting complete")

    except Exception as e:
        logger.error(f"Error formatting Excel sheet: {str(e)}")
        raise


def apply_excel_styling(
    ws, doctor_info: Dict, month: str, sessions: List[Dict], max_sessions: int
) -> None:
    """Apply styling to Excel worksheet"""
    logger.debug("Applying Excel styling...")

    # Color definitions
    weekend_fill = PatternFill(
        start_color="87CEEB", end_color="87CEEB", fill_type="solid"
    )
    total_monthly_hours_fill = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )
    session_header_fill = PatternFill(
        start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
    )

    # Font definitions
    header_font = Font(bold=True, size=11)
    total_monthly_hours_font = Font(bold=True, size=14)
    month_display_font = Font(bold=True, size=14)
    doctor_name_font = Font(bold=True, size=14)

    # Border definition
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Month display
    month_obj = datetime.strptime(month + "-01", "%Y-%m-%d")
    month_display = month_obj.strftime("%B %Y")

    # Calculate monthly total
    monthly_total = calculate_monthly_total_hours(sessions)

    # Apply header formatting
    apply_header_formatting(
        ws,
        month_display,
        doctor_info,
        monthly_total,
        max_sessions,
        month_display_font,
        doctor_name_font,
        total_monthly_hours_font,
        total_monthly_hours_fill,
        session_header_fill,
        header_font,
        thin_border,
    )


def apply_header_formatting(
    ws,
    month_display: str,
    doctor_info: Dict,
    monthly_total: float,
    max_sessions: int,
    month_display_font,
    doctor_name_font,
    total_monthly_hours_font,
    total_monthly_hours_fill,
    session_header_fill,
    header_font,
    thin_border,
) -> None:
    """Apply header formatting to Excel worksheet"""

    # DATE - merged vertically A1:B2
    ws.merge_cells("A1:B2")
    month_display_cell = ws["A1"]
    month_display_cell.value = month_display
    month_display_cell.font = month_display_font
    month_display_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply borders to merged range
    for row in range(1, 3):
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = thin_border

    # DOCTOR NAME - merged horizontally C1:D1
    ws.merge_cells("C1:D1")
    doctor_name_index_cell = ws["C1"]
    doctor_name_index_cell.value = "DOCTOR NAME:"
    doctor_name_index_cell.font = doctor_name_font

    for col in range(3, 5):
        ws.cell(row=1, column=col).border = thin_border

    # Doctor's full name in E1:F1
    ws.merge_cells("E1:F1")
    doctor_name_cell = ws["E1"]
    doctor_name_cell.value = format_doctor_first_last_name(doctor_info)
    doctor_name_cell.font = doctor_name_font
    doctor_name_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(5, 7):
        ws.cell(row=1, column=col).border = thin_border

    # Total Monthly Hours - merged horizontally G1:H1
    ws.merge_cells("G1:H1")
    total_monthly_cell = ws["G1"]
    total_monthly_cell.value = "Total Monthly Hours"
    total_monthly_cell.font = total_monthly_hours_font
    total_monthly_cell.fill = total_monthly_hours_fill
    total_monthly_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(7, 9):
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
        cell.fill = total_monthly_hours_fill

    # Monthly total value in I1
    ws["I1"] = f"{monthly_total:.1f}h"
    ws["I1"].font = total_monthly_hours_font
    ws["I1"].border = thin_border
    ws["I1"].fill = total_monthly_hours_fill

    # Session headers in row 2
    apply_session_headers(
        ws, max_sessions, session_header_fill, header_font, thin_border
    )

    # Column headers in row 3
    apply_column_headers(ws, max_sessions, header_font, thin_border)


def apply_session_headers(
    ws, max_sessions: int, session_header_fill, header_font, thin_border
) -> None:
    """Apply session headers to row 2"""
    col_start = 3  # Column C

    for i in range(max_sessions):
        session_start_col = col_start + (i * 4)  # 4 columns per session
        session_end_col = session_start_col + 3

        # Merge cells for session header
        ws.merge_cells(
            start_row=2,
            start_column=session_start_col,
            end_row=2,
            end_column=session_end_col,
        )

        cell = ws.cell(row=2, column=session_start_col)
        cell.value = f"SESSION {i + 1}"
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.fill = session_header_fill
        cell.border = thin_border

        # Apply styling to merged range
        for col in range(session_start_col, session_end_col + 1):
            range_cell = ws.cell(row=2, column=col)
            range_cell.border = thin_border
            range_cell.fill = session_header_fill


def apply_column_headers(ws, max_sessions: int, header_font, thin_border) -> None:
    """Apply column headers to row 3"""
    headers = generate_excel_column_headers(max_sessions)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.alignment = Alignment(horizontal="center")
        cell.font = header_font
        cell.border = thin_border


def add_excel_data_rows(
    ws, sessions: List[Dict], month_days: List[Dict], max_sessions: int
) -> None:
    """Add data rows to Excel worksheet"""
    logger.debug("Adding data rows to Excel worksheet")

    # Generate data rows
    data_rows = generate_excel_data_rows(sessions, month_days, max_sessions)

    # Weekend fill for highlighting
    weekend_fill = PatternFill(
        start_color="87CEEB", end_color="87CEEB", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add data rows starting from row 4
    for row_num, row_data in enumerate(data_rows, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = thin_border

            # Apply weekend highlighting
            day_info = month_days[row_num - 4]  # Adjust for 0-based index
            if day_info["is_weekend"]:
                cell.fill = weekend_fill


def auto_adjust_column_widths(ws, max_sessions: int) -> None:
    """Auto-adjust column widths for better readability"""
    logger.debug("Auto-adjusting Excel column widths")

    try:
        headers = generate_excel_column_headers(max_sessions)
        max_col = len(headers)

        for col_num in range(1, max_col + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)

            # Check all rows for this column
            for row_num in range(1, 50):  # Check reasonable number of rows
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    continue

            # Set column width
            adjusted_width = min(max(max_length + 2, 8), 25)  # Min 8, Max 25
            ws.column_dimensions[column_letter].width = adjusted_width

        logger.debug("Column width adjustment complete")

    except Exception as e:
        logger.warning(f"Error adjusting column widths: {str(e)}")


def generate_excel_file_bytes(
    team_leader_name: str, month: str, export_data: Dict[str, Any]
) -> io.BytesIO:
    """
    Generate the complete Excel file as BytesIO.

    Args:
        team_leader_name: Name of the team leader
        month: Month in YYYY-MM format
        export_data: Prepared export data

    Returns:
        BytesIO object containing the Excel file
    """
    logger.info(f"Generating Excel file for {team_leader_name}, {month}")

    try:
        # Create workbook
        wb = create_formatted_excel_workbook(team_leader_name, month, export_data)

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        logger.info("Excel file generation successful")
        return excel_file

    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        raise


def generate_export_filename(team_leader_name: str, month: str) -> str:
    """
    Generate export filename.

    Args:
        team_leader_name: Team leader name
        month: Month string

    Returns:
        Generated filename
    """
    month_obj = datetime.strptime(month + "-01", "%Y-%m-%d")
    filename = (
        f"{month_obj.strftime('%B_%Y')}_{team_leader_name.replace(' ', '_')}.xlsx"
    )
    logger.debug(f"Generated export filename: {filename}")
    return filename


# =============================================================================
# STREAMLIT UI COMPONENTS
# =============================================================================


def render_export_filters() -> Tuple[str, str]:
    """
    Render the filter section for export parameters.

    Returns:
        Tuple of (selected_team_leader, selected_month)
    """
    logger.debug("Rendering export filters")

    st.subheader("📋 Export Parameters")

    with st.container(border=True):
        col1, col2 = st.columns(2)

        with col1:
            # Team Leader selection
            tl_options = [POC_TEAM_LEADER["name"]]  # For now, only one TL
            selected_tl = st.selectbox(
                "Team Leader:", options=tl_options, key="export_tl_filter"
            )

        with col2:
            # Month selection
            selected_month = render_month_selection()

    logger.debug(f"Export filters - TL: {selected_tl}, Month: {selected_month}")
    return selected_tl, selected_month


def render_month_selection() -> str:
    """
    Render month selection for export.

    Returns:
        Selected month in YYYY-MM format
    """
    current_date = datetime.now()
    month_options = []

    # Generate last 7 months + next 6 months
    for i in range(-7, 7):  # -7 to +6 gives us 13 months total
        # More reliable month calculation
        year = current_date.year
        month = current_date.month + i

        # Handle year rollover
        while month > 12:
            month -= 12
            year += 1
        while month < 1:
            month += 12
            year -= 1

        # Create the date for this month
        month_date = datetime(year, month, 1)
        month_label = month_date.strftime("%B %Y")
        month_value = month_date.strftime("%Y-%m")
        month_options.append((month_label, month_value))

    # Sort by date to ensure proper order
    month_options.sort(key=lambda x: x[1])  # Sort by YYYY-MM value

    # Default to current month
    current_month_value = current_date.strftime("%Y-%m")
    month_labels = [opt[0] for opt in month_options]
    month_values = [opt[1] for opt in month_options]

    try:
        default_index = month_values.index(current_month_value)
    except ValueError:
        default_index = 12  # Middle of the range if current month not found

    selected_month_label = st.selectbox(
        "Month:",
        options=month_labels,
        index=default_index,
        key="export_month_filter",
    )

    selected_month = month_values[month_labels.index(selected_month_label)]
    return selected_month


def render_export_preview(export_data: Dict[str, Any], month: str) -> None:
    """
    Render a preview of what will be exported.

    Args:
        export_data: Prepared export data
        month: Month in YYYY-MM format
    """
    logger.debug(f"Rendering export preview for {len(export_data)} doctors")

    st.subheader("📊 Export Preview")

    with st.container(border=True):
        if not export_data:
            st.info("No data available for the selected month.")
            logger.warning("No export data available for preview")
            return

        # Summary stats
        render_export_summary_stats(export_data)

        # Doctor breakdown
        render_export_doctor_breakdown(export_data)


def render_export_summary_stats(export_data: Dict[str, Any]) -> None:
    """Render export summary statistics"""
    total_doctors = len(export_data)
    total_sessions = sum(len(data["sessions"]) for data in export_data.values())
    max_sessions_overall = (
        max(data["max_sessions"] for data in export_data.values()) if export_data else 0
    )

    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Doctors", total_doctors)
    with col2:
        st.metric("Total Sessions", total_sessions)
    with col3:
        st.metric("Max Sessions/Day", max_sessions_overall)


def render_export_doctor_breakdown(export_data: Dict[str, Any]) -> None:
    """Render doctor breakdown for export preview"""
    st.write("**Doctor Sessions Breakdown:**")

    for doctor_id, data in export_data.items():
        doctor_name = format_doctor_name(data["doctor_info"])
        session_count = len(data["sessions"])
        max_sessions = data["max_sessions"]

        if session_count > 0:
            monthly_total = calculate_monthly_total_hours(data["sessions"])
            st.write(
                f"• **{doctor_name}**: {session_count} sessions, "
                f"{monthly_total:.1f}h total, max {max_sessions} sessions/day"
            )
        else:
            st.write(f"• **{doctor_name}**: No sessions scheduled")


def render_export_generation_section(
    team_leader_name: str, month: str, export_data: Dict[str, Any]
) -> None:
    """
    Render the export button and handle file generation.

    Args:
        team_leader_name: Name of the team leader
        month: Month in YYYY-MM format
        export_data: Prepared export data
    """
    logger.debug("Rendering export generation section")

    st.subheader("📤 Generate Export")

    with st.container(border=True):
        if not export_data or not any(
            data["sessions"] for data in export_data.values()
        ):
            render_no_export_data_available()
            return

        # File name preview
        file_name = generate_export_filename(team_leader_name, month)
        st.write(f"**File name:** `{file_name}`")

        # Export button
        if st.button(
            "📊 Generate Excel File", type="primary", use_container_width=True
        ):
            handle_excel_generation(team_leader_name, month, export_data, file_name)


def render_no_export_data_available() -> None:
    """Render message when no export data is available"""
    st.warning("No session data available for export.")
    st.button("📊 Generate Excel File", disabled=True, help="No data to export")

    logger.warning("No export data available")


def handle_excel_generation(
    team_leader_name: str, month: str, export_data: Dict[str, Any], file_name: str
) -> None:
    """
    Handle Excel file generation and download.

    Args:
        team_leader_name: Team leader name
        month: Month string
        export_data: Export data
        file_name: Generated filename
    """
    logger.info(f"Handling Excel generation for {file_name}")

    try:
        with st.spinner("Generating Excel file..."):
            # Generate Excel file
            excel_file = generate_excel_file_bytes(team_leader_name, month, export_data)

            st.success("✅ Excel file generated successfully!")
            logger.info("Excel file generation successful")

            # Download button
            st.download_button(
                label="⬇️ Download Excel File",
                data=excel_file,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

            logger.info(f"Download button rendered for {file_name}")

    except Exception as e:
        error_msg = f"Error generating Excel file: {str(e)}"
        st.error(f"❌ {error_msg}")
        logger.error(error_msg)


def render_export_instructions() -> None:
    """Render export instructions and tips"""
    st.subheader("💡 Export Instructions")

    with st.container(border=True):
        st.write("**How to use the Excel export:**")
        st.write("1. Select the team leader and month you want to export")
        st.write("2. Review the preview to ensure correct data")
        st.write("3. Click 'Generate Excel File' to create the report")
        st.write("4. Download the file using the download button")

        st.write("**Excel file features:**")
        st.write("• One sheet per doctor")
        st.write("• Professional formatting with colors")
        st.write("• Weekend highlighting")
        st.write("• Monthly totals and daily summaries")
        st.write("• Combined scribe and patient number columns")


def render_export_page() -> None:
    """Main export page rendering function"""
    logger.info("Rendering export page")

    try:
        # Render header
        render_export_header()

        # Export instructions
        render_export_instructions()

        st.markdown("---")

        # Filters section
        selected_tl, selected_month = render_export_filters()

        st.markdown("---")

        # Prepare export data
        with st.spinner("Loading export data..."):
            export_data = prepare_export_data(
                POC_TEAM_LEADER["username"], selected_month
            )

        # Preview section
        render_export_preview(export_data, selected_month)

        st.markdown("---")

        # Export generation section
        render_export_generation_section(selected_tl, selected_month, export_data)

        # # Navigation back to main app
        # st.markdown("---")
        # render_navigation_section()

        logger.info("Export page rendering complete")

    except Exception as e:
        logger.error(f"Error rendering export page: {str(e)}")
        st.error("❌ An error occurred while loading the export page. Please refresh.")


def render_navigation_section() -> None:
    """Render navigation section"""
    st.subheader("🧭 Navigation")

    with st.container(border=True):
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**🏠 [Back to Main Schedule](http://localhost:8503)**")
            st.caption("Main scheduling interface")

        with col2:
            st.markdown("**📊 [View Analytics](http://localhost:8504)**")
            st.caption("Run: streamlit run src/ui/pages/audit.py --server.port 8504")


def validate_export_prerequisites() -> Tuple[bool, List[str]]:
    """
    Validate that all prerequisites for export are met.

    Returns:
        Tuple of (is_valid, list_of_issues)
    """
    logger.info("Validating export prerequisites")

    issues = []

    try:
        # Test Google Sheets connection
        if not validate_sheets_connection():
            issues.append("Google Sheets connection failed")

        # Check if we have any data
        stats = get_sheets_stats()
        if stats["total_submissions"] == 0:
            issues.append("No submission data available for export")

        # Check if we have doctor data
        if stats.get("total_doctors", 0) == 0:
            issues.append("No doctor data available")

        is_valid = len(issues) == 0

        if is_valid:
            logger.info("Export prerequisites validation successful")
        else:
            logger.warning(f"Export prerequisites validation failed: {issues}")

        return is_valid, issues

    except Exception as e:
        error_msg = f"Error validating export prerequisites: {str(e)}"
        logger.error(error_msg)
        return False, [error_msg]


def handle_export_error(error: Exception, context: str) -> None:
    """
    Handle export errors with proper logging and user feedback.

    Args:
        error: Exception that occurred
        context: Context where error occurred
    """
    error_msg = f"Export error in {context}: {str(error)}"
    logger.error(error_msg, exc_info=True)

    st.error(f"❌ {context} failed: {str(error)}")

    if is_debug_mode():
        st.exception(error)

    st.info("💡 Try refreshing the page or check your Google Sheets connection.")


def main() -> None:
    """Main export application entry point"""
    logger.info("Starting export page application")

    try:
        # Setup app
        setup_export_app()

        # Validate prerequisites
        is_valid, issues = validate_export_prerequisites()
        if not is_valid:
            st.error("❌ Export prerequisites not met:")
            for issue in issues:
                st.error(f"• {issue}")

            st.info("💡 Please resolve these issues before using the export feature.")
            return

        # Render export page
        render_export_page()

    except Exception as e:
        logger.critical(f"Critical error in export application: {str(e)}")
        handle_export_error(e, "Export application initialization")


# =============================================================================
# STREAMLIT ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    main()
